#!/usr/bin/env python3
"""Markdown 测试用例 → Excel 反向转换器

用法:
  python md_to_xlsx.py <input.md> [output.xlsx] [选项]

将本技能生成的结构化 Markdown 测试用例文件还原为 .xlsx 格式。
"""

import argparse
import json
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── 内置极简默认值 ──────────────────────────────────────────
_MINIMAL_COLUMN_ALIASES = {
    "用例编号": ["用例编号"], "模块": ["模块"], "用例标题": ["用例标题"],
    "优先级": ["优先级"], "前置条件": ["前置条件"], "测试步骤": ["测试步骤"],
    "预期结果": ["预期结果"], "备注": ["备注"], "设计方法": ["设计方法"],
    "需求编号": ["需求编号"],
}
_MINIMAL_PRIORITY_LEVELS = ["P0", "P1", "P2", "P3"]
_FIELDS_ORDER = ["用例编号", "模块", "用例标题", "优先级", "前置条件",
                 "测试步骤", "预期结果", "备注", "设计方法", "需求编号"]


def load_config(config_path: str | None = None) -> dict:
    paths_to_try = []
    if config_path:
        paths_to_try.append(Path(config_path))
    if not paths_to_try:
        print("ℹ️ 未指定配置文件，使用内置默认值")
        return {}
    for p in paths_to_try:
        if p.exists():
            with open(p, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            print(f"📋 已加载配置: {p}")
            return cfg
    print("ℹ️ 未找到配置文件，使用内置默认值")
    return {}


def _aliases(cfg: dict) -> dict[str, list[str]]:
    return cfg.get("column_aliases", _MINIMAL_COLUMN_ALIASES)


def _default_headers(cfg: dict) -> list[str]:
    """取每个标准字段名的第一个别名作为输出列头。"""
    aliases = _aliases(cfg)
    explicit = cfg.get("export_column_names")
    if explicit and isinstance(explicit, list) and len(explicit) == len(_FIELDS_ORDER):
        return explicit
    return [aliases.get(f, [f])[0] for f in _FIELDS_ORDER]


def _expand_module(cfg: dict) -> bool:
    return cfg.get("expand_module_name", False)


# ── Markdown 解析 ──────────────────────────────────────────

def parse_metadata(lines: list[str]) -> dict:
    meta = {"title": "", "source": "", "sheet": "Sheet1", "total": 0}
    for line in lines:
        m = re.match(r"^# (.+)$", line)
        if m:
            meta["title"] = m.group(1).strip()
        m = re.match(r"^>\s*来源：(.+?)\s*\|\s*Sheet[：:]\s*(.+?)\s*\|\s*总用例数[：:]\s*(\d+)", line)
        if m:
            meta["source"] = m.group(1).strip()
            meta["sheet"] = m.group(2).strip()
            meta["total"] = int(m.group(3))
    return meta


def parse_overview(lines: list[str]) -> tuple[dict[str, str], list[str]]:
    """解析模块概览表，返回 module_names 和 priority_levels。"""
    module_names = {}
    priority_levels = []
    in_overview = False
    for i, line in enumerate(lines):
        if line.strip() == "## 模块概览":
            in_overview = True
            continue
        if in_overview and line.startswith("---"):
            break
        if in_overview and line.startswith("|"):
            parts = [c.strip() for c in line.split("|")]
            parts = [p for p in parts if p]  # 去空
            # 表头行
            if parts[0] == "模块":
                # parts: [模块, 模块名称, 用例数, P0, P1, P2, P3, ...]
                if len(parts) > 3:
                    priority_levels = parts[3:]
                continue
            # 分割行跳过
            if parts[0].startswith("-"):
                continue
            # 合计行跳过
            if "合计" in parts[0]:
                continue
            # 数据行
            if len(parts) >= 2:
                code = parts[0].strip()
                name = parts[1].strip()
                if code and name:
                    module_names[code] = name
    return module_names, priority_levels if priority_levels else _MINIMAL_PRIORITY_LEVELS


def parse_modules(lines: list[str]) -> list[dict]:
    """按模块和用例编号分组，解析出 rows 列表。"""
    rows = []
    current_module = ""
    current_case = None
    current_field = None
    field_lines = []

    def _flush_field():
        nonlocal current_case, current_field, field_lines
        if current_case is None or current_field is None:
            return
        text = "\n".join(field_lines)
        if current_field == "前置条件":
            # 引用块：逐行去掉 > 前缀
            cleaned = []
            for ln in field_lines:
                cleaned.append(ln.removeprefix("> ").removeprefix(">"))
            text = "\n".join(cleaned)
        elif current_field == "预期结果":
            # 剥离自动添加的 - 前缀
            cleaned = []
            for ln in field_lines:
                s = ln.strip()
                if s.startswith("- ") and not re.match(r"^-\s*\d+\.", s):
                    cleaned.append(s[2:])
                else:
                    cleaned.append(ln)
            text = "\n".join(cleaned)
        current_case[current_field] = text
        field_lines = []
        current_field = None

    for line in lines:
        stripped = line.strip()

        # 模块标题行
        m = re.match(r"^## (\S+)(?:\s+(.*))?$", stripped)
        if m:
            _flush_field()
            code = m.group(1)
            # 跳过"模块概览"和非模块编号的章节（如"P0 新增用例"等补充章节）
            if code == "模块" or stripped == "## 模块概览":
                continue
            # 以优先级开头的章节（P0/P1/P2/P3/S1/S2/S3等）不是模块分组，跳过
            if re.match(r"^[PS]\d", code):
                continue
            current_module = code
            continue

        # 用例编号行
        m = re.match(r"^### (.+)$", stripped)
        if m:
            _flush_field()
            if current_case is not None:
                rows.append(current_case)
            current_case = {"用例编号": m.group(1).strip(), "模块": current_module}
            for f in _FIELDS_ORDER:
                if f not in current_case:
                    current_case[f] = ""
            current_field = None
            continue

        # 分割线
        if stripped == "---" or stripped == "***":
            _flush_field()
            if current_case is not None:
                rows.append(current_case)
                current_case = None
            current_field = None
            continue

        # 字段标记（支持粗体和纯文本两种格式）
        m = re.match(r"^\*\*(.+?)\*\*[：:]\s*(.*)$", stripped)
        if not m:
            m = re.match(r"^(.+?)[：:]\s*(.*)$", stripped)
        if m:
            _flush_field()
            field_name = m.group(1)
            rest = m.group(2).strip()
            # 字段名映射（Markdown 中用的中文标签 → 标准字段名）
            name_map = {"标题": "用例标题", "优先级": "优先级", "设计方法": "设计方法",
                        "需求编号": "需求编号", "备注": "备注"}
            std_name = name_map.get(field_name, field_name)
            if std_name in ("前置条件", "测试步骤", "预期结果"):
                # 多行字段，后续行在下方
                current_field = std_name
                field_lines = []
                # 如果冒号后有内容（一般不应有，但处理意外情况）
                if rest:
                    field_lines.append(rest)
            else:
                # 单行字段，直接赋值
                if current_case is not None:
                    current_case[std_name] = rest
                current_field = None
            continue

        # 引用块行（可出现在前置条件、测试步骤、预期结果中）
        if current_field in ("前置条件", "测试步骤", "预期结果") and stripped.startswith(">"):
            clean = stripped.removeprefix("> ").removeprefix(">")
            field_lines.append(clean)
            continue

        # 内容行（测试步骤、预期结果）
        if current_field in ("测试步骤", "预期结果"):
            if stripped:
                field_lines.append(stripped)
            elif not stripped:
                # 空行在步骤/预期中是分隔符，保留但不连续添加
                if field_lines and field_lines[-1] != "":
                    field_lines.append("")
            continue

        # 空行跳过
        if not stripped:
            continue

    # 处理最后一个用例
    _flush_field()
    if current_case is not None:
        rows.append(current_case)

    return rows


# ── xlsx 写入 ──────────────────────────────────────────

def write_xlsx(rows: list[dict], output_path: str, headers: list[str],
               sheet_name: str, module_names: dict, expand: bool = False):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 表头样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    thin_border = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, field in enumerate(_FIELDS_ORDER, 1):
            val = row_data.get(field, "")
            # 模块列可选展开
            if field == "模块" and expand and val in module_names:
                val = f"{val} {module_names[val]}"
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 自动列宽
    col_widths = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        for cell in row:
            if cell.value:
                length = max(len(str(line)) for line in str(cell.value).split("\n"))
                col = cell.column
                col_widths[col] = max(col_widths.get(col, 8), min(length + 2, 50))
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    print(f"✅ 已生成: {out}  ({len(rows)} 条用例)")


# ── 主函数 ──────────────────────────────────────────────

def convert(
    input_path: str,
    output_path: str | None = None,
    config_path: str | None = None,
    sheet_name: str | None = None,
) -> list[dict]:
    """主转换函数：md → xlsx"""

    cfg = load_config(config_path)
    headers = _default_headers(cfg)
    expand = _expand_module(cfg)

    md_path = Path(input_path)
    if not md_path.exists():
        raise FileNotFoundError(f"文件不存在: {input_path}")

    text = md_path.read_text(encoding="utf-8")
    lines = text.split("\n")

    # 1. 解析元信息
    meta = parse_metadata(lines)

    # 2. 解析概览表
    module_names, priority_levels = parse_overview(lines)

    # 3. 解析用例
    rows = parse_modules(lines)

    # 4. 校验
    if meta["total"] > 0 and len(rows) != meta["total"]:
        print(f"⚠️ 用例数不匹配: 元信息={meta['total']}, 实际解析={len(rows)}")

    empty_required = [r["用例编号"] for r in rows
                      if not r.get("用例编号") or not r.get("用例标题") or not r.get("优先级")]
    if empty_required:
        print(f"⚠️ 存在 {len(empty_required)} 条缺失必填字段的用例")

    # 5. 写入 xlsx
    out = Path(output_path) if output_path else md_path.with_suffix(".xlsx")
    sname = sheet_name or meta.get("sheet", "Sheet1")
    write_xlsx(rows, str(out), headers, sname, module_names, expand)

    return rows


def main():
    parser = argparse.ArgumentParser(description="Markdown 测试用例 → Excel 反向转换器")
    parser.add_argument("input", help="输入 .md 文件路径")
    parser.add_argument("output", nargs="?", default=None, help="输出 .xlsx 文件路径（默认同目录）")
    parser.add_argument("--config", default=None, help="配置文件路径（config.json）")
    parser.add_argument("--sheet-name", default=None, help="输出 Sheet 名称（默认从 md 元信息提取）")
    args = parser.parse_args()

    convert(args.input, args.output, config_path=args.config, sheet_name=args.sheet_name)


if __name__ == "__main__":
    main()